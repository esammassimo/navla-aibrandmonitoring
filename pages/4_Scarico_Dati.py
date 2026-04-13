"""pages/3_Scarico_Dati.py — Manual run, partial retry, scheduling UI."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from typing import Optional

import streamlit as st
from sqlalchemy import text

import pipeline as pl
from utils import (
    fetch_ai_questions,
    fetch_project_schedule,
    fetch_run_workers,
    fetch_runs,
    get_cookie_manager,
    get_engine,
    render_sidebar,
    require_login,
    set_schedule_active,
    upsert_project_schedule,
)

cookie_manager = get_cookie_manager()
require_login(cookie_manager)
render_sidebar(cookie_manager)

is_admin = st.session_state.get("role") == "admin"
project_id: Optional[str] = st.session_state.get("project_id")

st.title("Data Collection")

if not project_id:
    st.info("Select a project from the sidebar to get started.")
    st.stop()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
_ALL_LLMS = ["ChatGPT", "Claude", "Gemini", "Perplexity", "AI Overviews", "AI Mode"]
_ITERABLE_LLMS = ["ChatGPT", "Claude", "Gemini", "Perplexity"]  # LLMs that support multiple iterations


def _calc_next_run(frequency: str, day_of_week: int, day_of_month: int) -> datetime:
    today = date.today()
    if frequency == "weekly":
        days_ahead = (day_of_week - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        return datetime.combine(today + timedelta(days=days_ahead), datetime.min.time())
    else:  # biweekly or monthly
        month = today.month
        year = today.year
        if today.day >= day_of_month:
            month += 1
            if month > 12:
                month = 1
                year += 1
        return datetime(year, month, min(day_of_month, 28))



# ===========================================================================
# Export helpers
# ===========================================================================

def _build_export_xlsx(project_id: str, customer_id: str) -> bytes:
    """
    Build a Google Sheets-compatible .xlsx with all historical data for the
    given project/customer. Returns the file as bytes for st.download_button.

    Sheet structure (mirrors the Apps Script template):
      1. Readme         — static description
      2. Keyword        — keywords of the project
      3. AI Questions   — active + draft questions
      4. Brand - Apps Script   — v_brand_mentions_flat
      5. Fonti - Apps Script   — v_source_mentions_flat
      6. Risposte - Apps Script — v_ai_responses_flat
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from utils import run_query, FilterState, fetch_keywords, fetch_ai_questions

    # --- Palette (matches template) ---
    HDR_BG   = "FF29282C"   # dark charcoal
    HDR_FG   = "FFF0B910"   # gold/amber
    HDR_FONT = Font(bold=True, color=HDR_FG, name="Arial", size=10)
    HDR_FILL = PatternFill("solid", fgColor=HDR_BG)
    HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
    BODY_FONT = Font(name="Arial", size=10)
    BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)

    def _write_sheet(ws, headers: list[str], rows: list[list]):
        """Write header row + data rows with consistent styling."""
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = HDR_ALIGN

        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = BODY_FONT
                cell.alignment = BODY_ALIGN

        # Auto-width (capped at 80 chars)
        for c_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(c_idx)
            cell_lens = [len(str(h))] + [
                len(str(row[c_idx - 1])) if row[c_idx - 1] is not None else 0
                for row in rows[:200]
            ]
            ws.column_dimensions[col_letter].width = min(max(cell_lens) + 2, 80)

        # Freeze header row
        ws.freeze_panes = "A2"

    # --- Fetch data ---
    filters = FilterState(
        project_id=project_id,
        customer_id=customer_id,
        date_range=None,
        llms=(),
        clusters=(),
    )

    kw_df = run_query(
        "SELECT keyword, cluster, subcluster, search_volume "
        "FROM keywords WHERE project_id = %(pid)s ORDER BY cluster, keyword",
        {"pid": project_id},
    )
    q_df = run_query(
        "SELECT aq.question, k.keyword, k.cluster, k.subcluster, k.search_volume, "
        "aq.intent, aq.tone "
        "FROM ai_questions aq "
        "LEFT JOIN keywords k ON k.id = aq.keyword_id "
        "WHERE aq.project_id = %(pid)s ORDER BY k.cluster, aq.question",
        {"pid": project_id},
    )
    brand_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, brand, position, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_brand_mentions_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm, position",
        {"pid": project_id},
    )
    source_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, url, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_source_mentions_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm",
        {"pid": project_id},
    )
    response_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, response_text, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_ai_responses_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm",
        {"pid": project_id},
    )

    # --- Build workbook ---
    wb = Workbook()

    # 1. Keyword
    ws_kw = wb.active
    ws_kw.title = "Keyword"
    _write_sheet(
        ws_kw,
        headers=["Keyword", "CLUSTER", "SUBCLUSTER", "Volume"],
        rows=[
            [r.keyword, r.cluster, r.subcluster, r.search_volume]
            for r in kw_df.itertuples(index=False)
        ] if not kw_df.empty else [],
    )

    # 2. AI Questions
    ws_q = wb.create_sheet("AI Questions")
    _write_sheet(
        ws_q,
        headers=["AI Questions", "Keyword", "Cluster", "Subcluster", "Volume", "Intent", "Tone"],
        rows=[
            [r.question, r.keyword, r.cluster, r.subcluster, r.search_volume, r.intent, r.tone]
            for r in q_df.itertuples(index=False)
        ] if not q_df.empty else [],
    )

    # 3. Brand - Apps Script
    ws_brand = wb.create_sheet("Brand - Apps Script")
    _write_sheet(
        ws_brand,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "Brand", "Position", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.brand, r.position, r.intent, r.tone,
            ]
            for r in brand_df.itertuples(index=False)
        ] if not brand_df.empty else [],
    )

    # 4. Fonti - Apps Script
    ws_source = wb.create_sheet("Fonti - Apps Script")
    _write_sheet(
        ws_source,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "URL", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.url, r.intent, r.tone,
            ]
            for r in source_df.itertuples(index=False)
        ] if not source_df.empty else [],
    )

    # 5. Risposte - Apps Script
    ws_resp = wb.create_sheet("Risposte - Apps Script")
    # Response text can be long — wrap text and cap row height
    ws_resp.row_dimensions[1].height = 20
    _write_sheet(
        ws_resp,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "Risposta", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.response_text, r.intent, r.tone,
            ]
            for r in response_df.itertuples(index=False)
        ] if not response_df.empty else [],
    )
    # Risposta column (I) — wider and wrap text
    ws_resp.column_dimensions["I"].width = 80
    for row in ws_resp.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ===========================================================================
# SECTION 1 — Avvio manuale
# ===========================================================================
st.subheader("Manual run")

# Check active questions
q_df = fetch_ai_questions(project_id, status="active")
n_active = len(q_df)

if n_active == 0:
    st.warning(
        "No active questions for this project. "
        "Activate at least one question in the **Questions & Keywords** page before starting a run."
    )
else:
    st.caption(f"Active questions: **{n_active}**")

    with st.form("form_manual_run"):
        selected_llms = st.multiselect(
            "LLMs to query",
            options=_ALL_LLMS,
            default=_ALL_LLMS,
            help="Select at least one LLM.",
        )

        # Iterations slider — only meaningful for iterable LLMs
        iterable_selected = [l for l in selected_llms if l in _ITERABLE_LLMS] if selected_llms else []
        iterations = st.number_input(
            "Iterations per prompt",
            min_value=1,
            max_value=50,
            value=1,
            step=1,
            help=(
                "Number of times each question is sent to each supported LLM "
                f"({', '.join(_ITERABLE_LLMS)}). "
                "Iterations are sequential. AI Overviews and AI Mode are always queried once."
            ),
            disabled=not iterable_selected,
        )
        if iterable_selected and iterations > 1:
            st.caption(
                f"⚠️ With **{iterations} iteration(s)**, each question will be sent "
                f"**{iterations}×** to: {', '.join(iterable_selected)}. "
                f"Estimated total workers: **{n_active * (len(iterable_selected) * iterations + len([l for l in selected_llms if l not in _ITERABLE_LLMS]))}**"
            )

        run_btn = st.form_submit_button("▶ Start run", type="primary", disabled=(n_active == 0))

    if run_btn:
        if not selected_llms:
            st.error("Please select at least one LLM.")
        else:
            total_workers = n_active * len(selected_llms)
            progress_bar = st.progress(0, text="Starting…")
            status_text = st.empty()

            def _progress_cb(done: int, total: int) -> None:
                pct = done / total if total else 0
                progress_bar.progress(pct, text=f"Workers completed: {done}/{total}")
                status_text.caption(f"Running… {done}/{total}")

            try:
                run_id = pl.start_run(
                    project_id=project_id,
                    llms=selected_llms,
                    triggered_by="manual",
                    progress_callback=_progress_cb,
                    iterations=int(iterations),
                )
                progress_bar.progress(1.0, text="Run completed.")
                status_text.empty()
                st.success(f"Run completed successfully. ID: `{run_id}`")
                fetch_runs.clear()
            except ValueError as exc:
                st.error(str(exc))
            except Exception as exc:
                st.error(f"Error during run: {exc}")

# ===========================================================================
# SECTION 2 — Storico run
# ===========================================================================
st.divider()
st.subheader("Run history")

runs_df = fetch_runs(project_id)

if runs_df.empty:
    st.info("No runs executed for this project.")
else:
    display_cols = [
        "id", "started_at", "finished_at", "status",
        "triggered_by", "completed_questions", "total_questions",
    ]
    _col_config = {
        "id": st.column_config.TextColumn("Run ID"),
        "started_at": st.column_config.DatetimeColumn("Started", format="DD/MM/YY HH:mm"),
        "finished_at": st.column_config.DatetimeColumn("Finished", format="DD/MM/YY HH:mm"),
        "status": st.column_config.TextColumn("Status"),
        "triggered_by": st.column_config.TextColumn("Source"),
        "completed_questions": st.column_config.NumberColumn("Completed"),
        "total_questions": st.column_config.NumberColumn("Total"),
    }

    if is_admin:
        import pandas as pd

        sel_df = runs_df[display_cols].copy()
        sel_df.insert(0, "_sel", False)
        edited_runs = st.data_editor(
            sel_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "_sel": st.column_config.CheckboxColumn("Select", width="small"),
                **_col_config,
            },
            disabled=display_cols,
            key="runs_selector",
        )

        selected_ids = (
            runs_df.loc[edited_runs["_sel"].values, "id"].astype(str).tolist()
        )

        confirm_key = "confirm_del_runs"
        if selected_ids:
            if not st.session_state.get(confirm_key):
                if st.button(
                    f"🗑 Delete {len(selected_ids)} selected run(s)",
                    key="del_runs_btn",
                    type="primary",
                ):
                    st.session_state[confirm_key] = selected_ids
                    st.rerun()
            else:
                ids_to_del = st.session_state[confirm_key]
                st.error(
                    f"Deleting **{len(ids_to_del)} run(s)** will also remove all workers, "
                    "AI responses, brand mentions and source mentions. "
                    "This action cannot be undone."
                )
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Yes, delete", key="del_runs_confirm", type="primary"):
                        with get_engine().begin() as conn:
                            conn.execute(
                                text("DELETE FROM runs WHERE id::text = ANY(:ids)"),
                                {"ids": ids_to_del},
                            )
                        st.session_state.pop(confirm_key, None)
                        fetch_runs.clear()
                        st.success(f"Deleted {len(ids_to_del)} run(s).")
                        st.rerun()
                with c2:
                    if st.button("Cancel", key="del_runs_cancel"):
                        st.session_state.pop(confirm_key, None)
                        st.rerun()
    else:
        st.dataframe(
            runs_df[display_cols],
            use_container_width=True,
            hide_index=True,
            column_config=_col_config,
        )

# ===========================================================================
# SECTION 3 — Retry worker falliti
# ===========================================================================
st.divider()
st.subheader("Retry failed workers")

if runs_df.empty:
    st.info("No runs available.")
else:
    # Only runs that have partial or failed status
    retryable = runs_df[runs_df["status"].isin(["partial", "failed"])]

    if retryable.empty:
        st.success("No runs with failed workers.")
    else:
        run_opts = {
            f"{str(row['id'])[:8]}… — {row['status']} — {row['started_at']}": str(row["id"])
            for _, row in retryable.iterrows()
        }
        selected_run_label = st.selectbox(
            "Select run to retry", list(run_opts.keys()), key="retry_run_select"
        )
        selected_run_id = run_opts[selected_run_label]

        workers_df = fetch_run_workers(selected_run_id)
        failed_workers = workers_df[workers_df["status"] == "failed"] if not workers_df.empty else workers_df

        if failed_workers.empty:
            st.info("No failed workers in this run.")
        else:
            st.caption(f"Failed workers: **{len(failed_workers)}**")
            st.dataframe(
                failed_workers[["question", "llm", "attempt", "error", "started_at", "finished_at"]],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "question": st.column_config.TextColumn("Question"),
                    "llm": st.column_config.TextColumn("LLM"),
                    "attempt": st.column_config.NumberColumn("Attempt"),
                    "error": st.column_config.TextColumn("Error"),
                    "started_at": st.column_config.DatetimeColumn("Started", format="DD/MM/YY HH:mm"),
                    "finished_at": st.column_config.DatetimeColumn("Finished", format="DD/MM/YY HH:mm"),
                },
            )

            if st.button("🔄 Retry failed workers", type="primary", key="retry_btn"):
                total_retry = len(failed_workers)
                progress_bar_r = st.progress(0, text="Retrying…")
                status_text_r = st.empty()

                def _retry_cb(done: int, total: int) -> None:
                    pct = done / total if total else 0
                    progress_bar_r.progress(pct, text=f"Retry: {done}/{total}")
                    status_text_r.caption(f"Running… {done}/{total}")

                try:
                    pl.retry_failed_workers(
                        run_id=selected_run_id,
                        progress_callback=_retry_cb,
                    )
                    progress_bar_r.progress(1.0, text="Retry completed.")
                    status_text_r.empty()
                    st.success("Retry completed.")
                    fetch_runs.clear()
                    fetch_run_workers.clear()
                    st.rerun()
                except Exception as exc:
                    st.error(f"Error during retry: {exc}")

# ===========================================================================
# SECTION 4 — Scheduling  (admin only)
# ===========================================================================
if not is_admin:
    st.stop()

st.divider()
st.subheader("Automatic scheduling")

sched_df = fetch_project_schedule(project_id)

_FREQ_LABELS = {
    "weekly": "Weekly",
    "biweekly": "Biweekly",
    "monthly": "Monthly",
}
_DAYS_OF_WEEK = [
    "Monday", "Tuesday", "Wednesday", "Thursday",
    "Friday", "Saturday", "Sunday",
]

if not sched_df.empty:
    sched = sched_df.iloc[0]
    is_active: bool = bool(sched.get("is_active", False))

    col_status, col_toggle = st.columns([3, 1])
    with col_status:
        freq_label = _FREQ_LABELS.get(str(sched.get("frequency", "")), str(sched.get("frequency", "")))
        next_run = sched.get("next_run_at")
        next_str = (
            next_run.strftime("%d/%m/%Y") if hasattr(next_run, "strftime")
            else str(next_run) if next_run else "—"
        )
        status_label = "🟢 Active" if is_active else "🔴 Inactive"
        st.info(
            f"**Frequency:** {freq_label} &nbsp;|&nbsp; "
            f"**Next run:** {next_str} &nbsp;|&nbsp; "
            f"**Status:** {status_label}"
        )
    with col_toggle:
        toggle_label = "Deactivate" if is_active else "Activate"
        if st.button(toggle_label, key="toggle_sched"):
            set_schedule_active(project_id, not is_active)
            fetch_project_schedule.clear()
            st.rerun()

st.write("**Configure schedule**")

with st.form("form_schedule"):
    freq = st.selectbox(
        "Frequency",
        options=list(_FREQ_LABELS.keys()),
        format_func=lambda k: _FREQ_LABELS[k],
        index=0,
    )

    col_dow, col_dom = st.columns(2)
    with col_dow:
        day_of_week = st.selectbox(
            "Day of the week (weekly)",
            options=list(range(7)),
            format_func=lambda i: _DAYS_OF_WEEK[i],
            index=0,
            help="Used only for weekly frequency.",
        )
    with col_dom:
        day_of_month = st.number_input(
            "Day of the month (biweekly/monthly)",
            min_value=1, max_value=28,
            value=1,
            help="Used for biweekly and monthly frequency.",
        )

    sched_llms = st.multiselect(
        "LLM",
        options=_ALL_LLMS,
        default=_ALL_LLMS,
        help="LLMs to query in automatic runs.",
    )

    save_sched = st.form_submit_button("Save schedule", type="primary")

if save_sched:
    if not sched_llms:
        st.error("Please select at least one LLM.")
    else:
        next_run_at = _calc_next_run(freq, day_of_week, int(day_of_month))
        upsert_project_schedule(
            project_id,
            {
                "frequency": freq,
                "day_of_week": day_of_week,
                "day_of_month": int(day_of_month),
                "llms": sched_llms,
                "is_active": True,
                "next_run_at": next_run_at,
            },
        )
        fetch_project_schedule.clear()
        st.success(f"Schedule saved. Next run: **{next_run_at.strftime('%d/%m/%Y')}**")
        st.rerun()

# ===========================================================================
# SECTION 5 — Export Google Sheets
# ===========================================================================
st.divider()
st.subheader("Export data for Google Sheets")

customer_id: Optional[str] = st.session_state.get("customer_id")

st.caption(
    "Generate a .xlsx file compatible with the Google Sheets template containing "
    "all historical project data (keywords, questions, brands, sources, responses)."
)

col_exp1, col_exp2 = st.columns([3, 1])
with col_exp1:
    if runs_df.empty:
        st.info("No runs available for this project. Start at least one run before exporting.")
    else:
        n_runs = len(runs_df)
        completed_runs = int((runs_df["status"].isin(["completed", "partial"])).sum())
        st.caption(f"Total runs: **{n_runs}** &nbsp;|&nbsp; Completed/partial: **{completed_runs}**")

with col_exp2:
    export_disabled = runs_df.empty
    if st.button(
        "📥 Generate export",
        type="primary",
        disabled=export_disabled,
        use_container_width=True,
        key="btn_export_xlsx",
    ):
        with st.spinner("Generating file…"):
            try:
                xlsx_bytes = _build_export_xlsx(project_id, customer_id or "")
                st.session_state["export_xlsx_bytes"] = xlsx_bytes
                st.success("File generated. Click Download to save it.")
            except Exception as exc:
                st.error(f"Error during generation: {exc}")
                st.session_state.pop("export_xlsx_bytes", None)

if "export_xlsx_bytes" in st.session_state:
    from datetime import date as _date
    fname = f"AI_Brand_Monitor_{project_id[:8]}_{_date.today().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="⬇ Download .xlsx",
        data=st.session_state["export_xlsx_bytes"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        key="dl_export_xlsx",
    )

