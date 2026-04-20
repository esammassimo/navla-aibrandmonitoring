"""pages/1_Progetti.py — Wizard 4-step: crea un nuovo progetto (admin only)."""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from typing import Optional

import pandas as pd
import requests
import streamlit as st

from utils import (
    create_project,
    get_cookie_manager,
    get_engine,
    insert_keywords,
    insert_ai_questions,
    render_sidebar,
    require_login,
    upsert_project_brands,
    upsert_project_schedule,
    run_query,
)
from sqlalchemy import text

cookie_manager = get_cookie_manager()
require_login(cookie_manager)
render_sidebar(cookie_manager)

if st.session_state.get("role") != "admin":
    st.error("Accesso riservato agli amministratori.")
    st.stop()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
LANGUAGES = {
    "Italiano": "it", "English": "en", "Deutsch": "de",
    "Français": "fr", "Español": "es", "Português": "pt",
    "Nederlands": "nl", "Polski": "pl",
}
LANG_DEFAULT_COUNTRY = {
    "it": "it", "en": "us", "de": "de",
    "fr": "fr", "es": "es", "pt": "br",
    "nl": "nl", "pl": "pl",
}
COUNTRIES = [
    "us", "it", "de", "fr", "es", "pt", "br", "gb",
    "nl", "pl", "ch", "at", "be", "mx", "ar", "co",
    "cl", "au", "ca", "jp", "in", "za",
]
DOW_LABELS = {
    "Lunedì": 0, "Martedì": 1, "Mercoledì": 2, "Giovedì": 3,
    "Venerdì": 4, "Sabato": 5, "Domenica": 6,
}
FREQ_LABELS = {"Settimanale": "weekly", "Bisettimanale": "biweekly", "Mensile": "monthly"}
LLM_OPTIONS = ["ChatGPT", "Claude", "Gemini", "Perplexity", "AI Overviews", "AI Mode"]

# ---------------------------------------------------------------------------
# Wizard state helpers  (prefix: wiz1_)
# ---------------------------------------------------------------------------
def _get(key, default=None):
    k = f"wiz1_{key}"
    if k not in st.session_state:
        st.session_state[k] = default
    return st.session_state[k]


def _set(key, value):
    st.session_state[f"wiz1_{key}"] = value


def _reset():
    for k in list(st.session_state.keys()):
        if k.startswith("wiz1_"):
            del st.session_state[k]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def _parse_file(uploaded) -> Optional[pd.DataFrame]:
    """Parse uploaded CSV or Excel file into a DataFrame."""
    try:
        if uploaded.name.lower().endswith(".csv"):
            return pd.read_csv(uploaded)
        return pd.read_excel(uploaded)
    except Exception as exc:
        st.error(f"Errore lettura file: {exc}")
        return None


def _parse_brands(raw: str) -> list[str]:
    """Split brand string on newlines and commas, deduplicate, strip."""
    if not raw.strip():
        return []
    parts = re.split(r"[\n,]+", raw)
    seen: set[str] = set()
    result = []
    for p in parts:
        p = p.strip()
        if p and p.lower() not in seen:
            seen.add(p.lower())
            result.append(p)
    return result


def _fetch_paa(keywords: list[str], language: str, country: str) -> pd.DataFrame:
    """Fetch People Also Ask via SerpApi for a list of keywords (max 4 per keyword)."""
    api_key = st.secrets.get("api_keys", {}).get("serpapi", "")
    if not api_key:
        st.error("SerpApi key non configurata in `.streamlit/secrets.toml`.")
        return pd.DataFrame()

    rows: list[dict] = []
    progress = st.progress(0)
    total = len(keywords)

    for i, kw in enumerate(keywords):
        try:
            resp = requests.get(
                "https://serpapi.com/search.json",
                params={
                    "engine": "google",
                    "q": kw,
                    "api_key": api_key,
                    "gl": country,
                    "hl": language,
                },
                timeout=15,
            )
            if resp.status_code != 200:
                st.warning(f"SerpApi errore {resp.status_code} per '{kw}': {resp.text[:200]}")
                progress.progress((i + 1) / total)
                continue

            seen: set[str] = set()
            count = 0
            for item in resp.json().get("related_questions", []):
                q = item.get("question", "").strip()
                if q and q not in seen:
                    seen.add(q)
                    rows.append({"keyword": kw, "question": q})
                    count += 1
                    if count >= 4:
                        break
        except Exception as exc:
            st.warning(f"Errore SerpApi per '{kw}': {exc}")

        progress.progress((i + 1) / total)

    progress.empty()
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["keyword", "question"])


def _save_percorso_a(project_id: str, df: pd.DataFrame) -> None:
    """Save keywords + questions from a full import file (Percorso A)."""
    # Normalise column names
    df = df.rename(columns={"volume": "search_volume"})

    # 1. Insert unique keywords
    kw_df = (
        df[["keyword"] + [c for c in ("cluster", "subcluster", "search_volume") if c in df.columns]]
        .drop_duplicates("keyword")
    )
    insert_keywords(project_id, kw_df.to_dict("records"))

    # 2. Build keyword → id map (bypass cache with run_query)
    kw_saved = run_query(
        "SELECT id, keyword FROM keywords WHERE project_id = %(pid)s",
        {"pid": project_id},
    )
    kw_id_map: dict[str, str] = dict(zip(kw_saved["keyword"], kw_saved["id"].astype(str)))

    # 3. Insert questions
    q_rows = []
    for _, row in df.iterrows():
        q = str(row.get("question", "")).strip()
        if q:
            q_rows.append({
                "keyword_id": kw_id_map.get(str(row.get("keyword", ""))),
                "question": q,
                "intent": row.get("intent") if pd.notna(row.get("intent")) else None,
                "tone": row.get("tone") if pd.notna(row.get("tone")) else None,
                "source": "csv_import",
                "status": "active",
            })
    insert_ai_questions(project_id, q_rows)


def _save_percorso_b(
    project_id: str,
    kw_df: pd.DataFrame,
    paa_df: Optional[pd.DataFrame],
) -> None:
    """Save keywords + selected PAA questions (Percorso B)."""
    # 1. Insert keywords
    kw_df = kw_df.rename(columns={"volume": "search_volume"})
    insert_keywords(project_id, kw_df.to_dict("records"))

    if paa_df is None or paa_df.empty:
        return

    # 2. Filter to selected rows only
    if "seleziona" in paa_df.columns:
        paa_df = paa_df[paa_df["seleziona"] == True]  # noqa: E712

    if paa_df.empty:
        return

    # 3. Build keyword → id map
    kw_saved = run_query(
        "SELECT id, keyword FROM keywords WHERE project_id = %(pid)s",
        {"pid": project_id},
    )
    kw_id_map: dict[str, str] = dict(zip(kw_saved["keyword"], kw_saved["id"].astype(str)))

    # 4. Insert questions
    q_rows = []
    for _, row in paa_df.iterrows():
        q = str(row.get("question", "")).strip()
        if q:
            q_rows.append({
                "keyword_id": kw_id_map.get(str(row.get("keyword", ""))),
                "question": q,
                "source": "serpapi_paa",
                "status": "active",
            })
    insert_ai_questions(project_id, q_rows)


def _import_from_excel(project_id: str, uploaded_file) -> dict:
    """
    Import a full Compass-format Excel file into an existing project.
    Creates: keywords, ai_questions, and historical runs with
    ai_responses, brand_mentions, source_mentions.
    Returns a summary dict with counts.
    """
    import math
    from openpyxl import load_workbook

    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    summary = {
        "keywords": 0, "questions": 0, "runs": 0,
        "responses": 0, "brands": 0, "sources": 0, "errors": [],
    }

    # --- 1. Keywords ---
    if "Keyword" in wb.sheetnames:
        ws = wb["Keyword"]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if len(rows) > 1:
            kw_rows = []
            for r in rows[1:]:
                kw = str(r[0]).strip() if r[0] else None
                if not kw:
                    continue
                def _clean(v):
                    if v is None: return None
                    if isinstance(v, float) and math.isnan(v): return None
                    return str(v).strip() or None
                kw_rows.append({
                    "keyword":    kw,
                    "cluster":    _clean(r[1] if len(r) > 1 else None),
                    "subcluster": _clean(r[2] if len(r) > 2 else None),
                    "search_volume": int(r[3]) if len(r) > 3 and r[3] and not (isinstance(r[3], float) and math.isnan(r[3])) else None,
                })
            if kw_rows:
                insert_keywords(project_id, kw_rows)
                summary["keywords"] = len(kw_rows)

    # --- 2. AI Questions (match keyword by text) ---
    if "AI Questions" in wb.sheetnames:
        ws = wb["AI Questions"]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        # Reload keyword IDs
        kw_df = run_query(
            "SELECT id, keyword FROM keywords WHERE project_id = %(pid)s",
            {"pid": project_id},
        )
        kw_map = {str(r["keyword"]).strip().lower(): str(r["id"]) for _, r in kw_df.iterrows()}

        q_rows = []
        for r in rows[1:]:
            question = str(r[0]).strip() if r[0] else None
            if not question:
                continue
            kw_text = str(r[1]).strip().lower() if len(r) > 1 and r[1] else ""
            kw_id = kw_map.get(kw_text)
            intent = str(r[5]).strip() if len(r) > 5 and r[5] else None
            tone   = str(r[6]).strip() if len(r) > 6 and r[6] else None
            q_rows.append({
                "question":   question,
                "keyword_id": kw_id,
                "intent":     intent,
                "tone":       tone,
                "source":     "csv_import",
                "status":     "active",
            })
        if q_rows:
            insert_ai_questions(project_id, q_rows)
            summary["questions"] = len(q_rows)

    # --- 3. Historical data: one run per date ---
    risposta_sheet = "Risposte - Apps Script"
    brand_sheet    = "Brand - Apps Script"
    fonti_sheet    = "Fonti - Apps Script"

    missing = [s for s in [risposta_sheet, brand_sheet, fonti_sheet] if s not in wb.sheetnames]
    if missing:
        summary["errors"].append(f"Missing sheets: {missing}")
        return summary

    def _load_sheet(ws_name):
        ws = wb[ws_name]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if len(rows) < 2:
            return pd.DataFrame()
        ncols = len(rows[0])
        data_rows = [r[:ncols] for r in rows[1:]]
        return pd.DataFrame(data_rows, columns=rows[0])

    df_resp   = _load_sheet(risposta_sheet)
    df_brand  = _load_sheet(brand_sheet)
    df_source = _load_sheet(fonti_sheet)

    if df_resp.empty:
        return summary

    # Reload question IDs
    q_df = run_query(
        "SELECT id, question FROM ai_questions WHERE project_id = %(pid)s",
        {"pid": project_id},
    )
    q_map = {str(r["question"]).strip().lower(): str(r["id"]) for _, r in q_df.iterrows()}

    # Group responses by date — one run per date
    df_resp["_date"] = pd.to_datetime(df_resp["Data"]).dt.date
    df_brand["_date"]  = pd.to_datetime(df_brand["Data"]).dt.date
    df_source["_date"] = pd.to_datetime(df_source["Data"]).dt.date

    for run_date, resp_group in df_resp.groupby("_date"):
        llms_in_run = resp_group["LLM"].dropna().unique().tolist()
        n_questions = len(resp_group)

        # Create run record
        with get_engine().begin() as conn:
            row = conn.execute(
                text(
                    "INSERT INTO runs "
                    "(project_id, started_at, finished_at, status, triggered_by, llms, "
                    " total_questions, completed_questions) "
                    "VALUES (:pid, :started, :finished, 'completed', 'manual', :llms, "
                    "        :total, :total) "
                    "RETURNING id"
                ),
                {
                    "pid":      project_id,
                    "started":  datetime.combine(run_date, datetime.min.time()),
                    "finished": datetime.combine(run_date, datetime.min.time()),
                    "llms":     llms_in_run,
                    "total":    n_questions,
                },
            ).fetchone()
            run_id = str(row[0])

        summary["runs"] += 1

        # Brand and source data for this date
        b_group = df_brand[df_brand["_date"] == run_date] if not df_brand.empty else pd.DataFrame()
        s_group = df_source[df_source["_date"] == run_date] if not df_source.empty else pd.DataFrame()

        # Insert ai_responses + brand_mentions + source_mentions
        for _, resp_row in resp_group.iterrows():
            question_text = str(resp_row.get("AI Questions", "")).strip()
            llm    = str(resp_row.get("LLM", "")).strip()
            model  = str(resp_row.get("Model", "")).strip()
            resp_text = str(resp_row.get("Risposta", "")).strip() or None
            q_id = q_map.get(question_text.lower())

            with get_engine().begin() as conn:
                ar_row = conn.execute(
                    text(
                        "INSERT INTO ai_responses "
                        "(run_id, ai_question_id, llm, model, response_text, run_date) "
                        "VALUES (:rid, :qid, :llm, :model, :text, :rdate) "
                        "RETURNING id"
                    ),
                    {
                        "rid":   run_id,
                        "qid":   q_id,
                        "llm":   llm,
                        "model": model,
                        "text":  resp_text,
                        "rdate": run_date,
                    },
                ).fetchone()
                response_id = str(ar_row[0])
            summary["responses"] += 1

            # Brand mentions for this response
            b_match = b_group[
                (b_group["AI Questions"].astype(str).str.strip() == question_text) &
                (b_group["LLM"].astype(str).str.strip() == llm)
            ] if not b_group.empty else pd.DataFrame()

            if not b_match.empty:
                with get_engine().begin() as conn:
                    for _, br in b_match.iterrows():
                        brand_name = str(br.get("Brand", "")).strip()
                        position   = br.get("Position")
                        if not brand_name:
                            continue
                        pos_int = int(position) if position and not (isinstance(position, float) and math.isnan(position)) else None
                        conn.execute(
                            text("INSERT INTO brand_mentions (ai_response_id, brand_name, position) "
                                 "VALUES (:rid, :brand, :pos)"),
                            {"rid": response_id, "brand": brand_name, "pos": pos_int},
                        )
                        summary["brands"] += 1

            # Source mentions for this response
            s_match = s_group[
                (s_group["AI Questions"].astype(str).str.strip() == question_text) &
                (s_group["LLM"].astype(str).str.strip() == llm)
            ] if not s_group.empty else pd.DataFrame()

            if not s_match.empty:
                with get_engine().begin() as conn:
                    for _, sr in s_match.iterrows():
                        url = str(sr.get("URL", "")).strip()
                        if not url:
                            continue
                        conn.execute(
                            text("INSERT INTO source_mentions (ai_response_id, url) "
                                 "VALUES (:rid, :url)"),
                            {"rid": response_id, "url": url},
                        )
                        summary["sources"] += 1

    return summary


def _calc_next_run(
    frequency: str,
    day_of_week: Optional[int],
    day_of_month: Optional[int],
) -> datetime:
    """Calculate the next scheduled run datetime."""
    today = date.today()
    if frequency in ("weekly", "biweekly"):
        days_ahead = (day_of_week - today.weekday()) % 7 or 7
        next_date = today + timedelta(days=days_ahead)
        if frequency == "biweekly":
            next_date += timedelta(weeks=1)
    else:  # monthly
        dom = min(day_of_month or 1, 28)
        if today.day < dom:
            next_date = today.replace(day=dom)
        elif today.month == 12:
            next_date = date(today.year + 1, 1, dom)
        else:
            next_date = date(today.year, today.month + 1, dom)
    return datetime.combine(next_date, datetime.min.time())


def _finalize() -> None:
    """Move to the completion step and clear caches."""
    _set("step", 5)
    st.cache_data.clear()
    st.rerun()


# ---------------------------------------------------------------------------
# Initialise step
# ---------------------------------------------------------------------------
if _get("step") is None:
    _set("step", 1)

step = _get("step")

# ---------------------------------------------------------------------------
# Step indicator
# ---------------------------------------------------------------------------
st.title("Nuovo Progetto")

STEP_LABELS = ["Dati progetto", "Keyword & Domande", "Competitor", "Scheduling"]
cols = st.columns(4)
for i, label in enumerate(STEP_LABELS):
    with cols[i]:
        if i + 1 < step:
            st.success(f"✓ {label}")
        elif i + 1 == step:
            st.info(f"**{i + 1}. {label}**")
        else:
            st.caption(f"{i + 1}. {label}")

st.divider()

# ===========================================================================
# STEP 1 — Dati progetto
# ===========================================================================
if step == 1:
    st.subheader("Step 1 — Dati del progetto")
    customer_id = st.session_state.get("customer_id")

    if not customer_id:
        st.error("Nessun cliente selezionato. Torna alla pagina **Clienti** e crea/seleziona un cliente.")
        st.stop()

    with st.form("step1_form"):
        name = st.text_input("Nome progetto", value=_get("name", ""))
        lang_label = st.selectbox("Lingua", list(LANGUAGES.keys()))
        lang_code = LANGUAGES[lang_label]

        default_country = LANG_DEFAULT_COUNTRY.get(lang_code, "us")
        country_idx = COUNTRIES.index(default_country) if default_country in COUNTRIES else 0
        country = st.selectbox("Paese", COUNTRIES, index=country_idx)

        submitted = st.form_submit_button("Avanti →", type="primary")

    if submitted:
        if not name.strip():
            st.error("Il nome del progetto è obbligatorio.")
        else:
            with st.spinner("Creazione progetto…"):
                project_id = create_project(customer_id, name.strip(), lang_code, country)
            _set("project_id", project_id)
            _set("name", name.strip())
            _set("language", lang_code)
            _set("country", country)
            _set("step", 2)
            st.cache_data.clear()
            st.rerun()

# ===========================================================================
# STEP 2 — Keyword & Domande
# ===========================================================================
elif step == 2:
    st.subheader("Step 2 — Keyword e Domande")
    project_id: str = _get("project_id")
    language: str = _get("language", "en")
    country: str = _get("country", "us")

    path = st.radio(
        "Cosa hai a disposizione?",
        options=[
            "A — Ho keyword e domande (file completo)",
            "B — Ho solo keyword (recupero PAA da SerpApi)",
            "C — Non ho nulla, aggiungo in seguito",
            "D — Importa da file Excel (formato Compass)",
        ],
        key="wiz1_path",
        horizontal=True,
    )

    # ---- PERCORSO A ----
    if path.startswith("A"):
        st.caption(
            "Colonne attese: `keyword`, `cluster` *(opz.)*, `subcluster` *(opz.)*, "
            "`volume` *(opz.)*, `question`, `intent` *(opz.)*, `tone` *(opz.)*"
        )
        uploaded = st.file_uploader(
            "Carica file CSV o Excel", type=["csv", "xlsx", "xls"], key="wiz1_upload_a"
        )
        if uploaded:
            df = _parse_file(uploaded)
            if df is not None:
                missing = [c for c in ("keyword", "question") if c not in df.columns]
                if missing:
                    st.error(f"Colonne obbligatorie mancanti: {missing}")
                else:
                    st.success(f"{len(df)} righe caricate. Anteprima:")
                    st.dataframe(df.head(20), use_container_width=True, hide_index=True)
                    _set("file_a_df", df)

        c1, c2 = st.columns([1, 6])
        with c1:
            if st.button("← Indietro", key="a_back"):
                _set("step", 1)
                st.rerun()
        with c2:
            if st.button("Salva e avanti →", type="primary", key="a_next",
                         disabled=_get("file_a_df") is None):
                with st.spinner("Salvataggio keyword e domande…"):
                    _save_percorso_a(project_id, _get("file_a_df"))
                _set("step", 3)
                st.rerun()

    # ---- PERCORSO B ----
    elif path.startswith("B"):
        st.caption("Colonne attese: `keyword`, `cluster` *(opz.)*, `subcluster` *(opz.)*, `volume` *(opz.)*")
        uploaded = st.file_uploader(
            "Carica file CSV o Excel (keyword)", type=["csv", "xlsx", "xls"], key="wiz1_upload_b"
        )
        if uploaded:
            df = _parse_file(uploaded)
            if df is not None:
                if "keyword" not in df.columns:
                    st.error("Colonna `keyword` mancante.")
                else:
                    st.success(f"{len(df)} keyword caricate.")
                    st.dataframe(df.head(20), use_container_width=True, hide_index=True)
                    _set("file_b_df", df)

        if _get("file_b_df") is not None:
            if st.button("🔍 Recupera People Also Ask da SerpApi", key="b_paa"):
                kws = _get("file_b_df")["keyword"].dropna().unique().tolist()
                paa_df = _fetch_paa(kws, language, country)
                if not paa_df.empty:
                    paa_df.insert(0, "seleziona", False)
                    _set("paa_df", paa_df)
                    st.success(f"Trovate {len(paa_df)} domande PAA.")
                else:
                    st.warning("Nessuna PAA trovata per le keyword caricate.")

        if _get("paa_df") is not None:
            paa = _get("paa_df")
            st.subheader(f"Domande PAA ({len(paa)}) — seleziona quelle da importare")
            edited = st.data_editor(
                paa,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "seleziona": st.column_config.CheckboxColumn("Importa", default=False),
                    "keyword": st.column_config.TextColumn("Keyword"),
                    "question": st.column_config.TextColumn("Domanda PAA"),
                },
                key="wiz1_paa_editor",
            )
            n_sel = int(edited["seleziona"].sum())
            st.caption(f"{n_sel} domande selezionate")
            _set("paa_df_edited", edited)

        c1, c2 = st.columns([1, 6])
        with c1:
            if st.button("← Indietro", key="b_back"):
                _set("step", 1)
                st.rerun()
        with c2:
            if st.button("Salva e avanti →", type="primary", key="b_next",
                         disabled=_get("file_b_df") is None):
                _edited = _get("paa_df_edited")
                edited = _edited if _edited is not None else _get("paa_df")
                with st.spinner("Salvataggio keyword e domande…"):
                    _save_percorso_b(project_id, _get("file_b_df"), edited)
                _set("step", 3)
                st.rerun()

    # ---- PERCORSO D ----
    elif path.startswith("D"):
        st.caption(
            "Importa un file Excel nel **formato Compass** con i fogli: "
            "`Keyword`, `AI Questions`, `Brand - Apps Script`, "
            "`Fonti - Apps Script`, `Risposte - Apps Script`."
        )
        uploaded_d = st.file_uploader(
            "Carica file Excel (.xlsx)", type=["xlsx"], key="wiz1_upload_d"
        )

        if uploaded_d:
            # Preview each sheet
            import openpyxl as _oxl
            wb_prev = _oxl.load_workbook(uploaded_d, read_only=True, data_only=True)
            uploaded_d.seek(0)  # reset for later use

            tabs = st.tabs(["Keyword", "AI Questions", "Brand", "Fonti", "Risposte"])
            sheet_map = {
                "Keyword":              "Keyword",
                "AI Questions":         "AI Questions",
                "Brand":                "Brand - Apps Script",
                "Fonti":                "Fonti - Apps Script",
                "Risposte":             "Risposte - Apps Script",
            }
            counts = {}
            for tab, (tab_label, sheet_name) in zip(tabs, sheet_map.items()):
                with tab:
                    if sheet_name in wb_prev.sheetnames:
                        ws_p = wb_prev[sheet_name]
                        rows_p = [r for r in ws_p.iter_rows(values_only=True)
                                  if any(c is not None for c in r)]
                        n_rows = max(0, len(rows_p) - 1)
                        counts[tab_label] = n_rows
                        st.caption(f"**{n_rows}** righe")
                        if rows_p:
                            preview_df = pd.DataFrame(
                                rows_p[1:min(6, len(rows_p))],
                                columns=[str(c) for c in rows_p[0][:8]]
                            )
                            st.dataframe(preview_df, use_container_width=True, hide_index=True)
                    else:
                        st.warning(f"Foglio '{sheet_name}' non trovato nel file.")

            _set("import_file_d", uploaded_d)

            total_rows = sum(counts.values())
            st.info(
                f"📊 **Riepilogo:** "
                f"{counts.get('Keyword', 0)} keyword · "
                f"{counts.get('AI Questions', 0)} domande · "
                f"{counts.get('Risposte', 0)} risposte storiche · "
                f"{counts.get('Brand', 0)} brand mentions · "
                f"{counts.get('Fonti', 0)} source mentions"
            )

        c1, c2 = st.columns([1, 6])
        with c1:
            if st.button("← Indietro", key="d_back"):
                _set("step", 1)
                st.rerun()
        with c2:
            if st.button("Importa tutto →", type="primary", key="d_next",
                         disabled=_get("import_file_d") is None):
                f = _get("import_file_d")
                f.seek(0)
                with st.spinner("Importazione in corso… potrebbe richiedere qualche minuto."):
                    result = _import_from_excel(project_id, f)
                if result["errors"]:
                    st.error(f"Errori durante l'importazione: {result['errors']}")
                else:
                    st.success(
                        f"✅ Importazione completata: "
                        f"**{result['keywords']}** keyword · "
                        f"**{result['questions']}** domande · "
                        f"**{result['runs']}** run storici · "
                        f"**{result['responses']}** risposte · "
                        f"**{result['brands']}** brand mentions · "
                        f"**{result['sources']}** source mentions"
                    )
                    st.cache_data.clear()
                _set("step", 3)
                st.rerun()

    # ---- PERCORSO C ----
    else:
        st.info(
            "Il progetto verrà creato senza keyword o domande. "
            "Potrai aggiungerle in seguito dalla pagina **Domande e Keyword**."
        )
        c1, c2 = st.columns([1, 6])
        with c1:
            if st.button("← Indietro", key="c_back"):
                _set("step", 1)
                st.rerun()
        with c2:
            if st.button("Avanti →", type="primary", key="c_next"):
                _set("step", 3)
                st.rerun()

# ===========================================================================
# STEP 3 — Competitor brands
# ===========================================================================
elif step == 3:
    st.subheader("Step 3 — Brand e Competitor (opzionale)")
    project_id = _get("project_id")

    st.caption(
        "Inserisci i brand da monitorare (uno per riga o separati da virgola). "
        "Per ciascuno indica se è un competitor o un tuo brand proprio."
    )

    raw_input = st.text_area(
        "Brand (uno per riga o separati da virgola)",
        value=_get("brands_raw", ""),
        height=140,
        key="wiz1_brands_text",
        placeholder="Locauto\nHertz\nSixt",
    )
    _set("brands_raw", raw_input)

    parsed = _parse_brands(raw_input)
    brand_competitor: dict[str, bool] = {}
    brand_own: dict[str, bool] = {}

    if parsed:
        st.write("**Configura i brand:**")
        header_cols = st.columns([3, 1, 1])
        header_cols[0].markdown("**Brand**")
        header_cols[1].markdown("**Competitor?**")
        header_cols[2].markdown("**Brand proprio?**")
        for brand in parsed:
            safe = re.sub(r"[^a-z0-9]", "_", brand.lower())
            row = st.columns([3, 1, 1])
            row[0].write(brand)
            brand_competitor[brand] = row[1].checkbox(
                "Competitor", key=f"wiz1_comp_{safe}", label_visibility="collapsed"
            )
            brand_own[brand] = row[2].checkbox(
                "Brand proprio", key=f"wiz1_own_{safe}", label_visibility="collapsed"
            )

    c1, c2, c3 = st.columns([1, 1, 5])
    with c1:
        if st.button("← Indietro", key="s3_back"):
            _set("step", 2)
            st.rerun()
    with c2:
        if st.button("Salta →", key="s3_skip"):
            _set("step", 4)
            st.rerun()
    with c3:
        if st.button("Salva e avanti →", type="primary", key="s3_next",
                     disabled=not parsed):
            conflicting = [b for b in parsed if brand_competitor.get(b) and brand_own.get(b)]
            if conflicting:
                st.error(
                    f"Un brand non può essere sia competitor che brand proprio: "
                    f"**{', '.join(conflicting)}**"
                )
            else:
                brands = [
                    {
                        "brand_name": b,
                        "is_competitor": brand_competitor.get(b, False),
                        "is_own_brand": brand_own.get(b, False),
                    }
                    for b in parsed
                ]
                upsert_project_brands(project_id, brands)
            _set("step", 4)
            st.rerun()

# ===========================================================================
# STEP 4 — Schedule automatico
# ===========================================================================
elif step == 4:
    st.subheader("Step 4 — Scheduling automatico (opzionale)")
    project_id = _get("project_id")

    freq_label = st.selectbox("Frequenza", list(FREQ_LABELS.keys()), key="wiz1_freq")
    freq = FREQ_LABELS[freq_label]

    day_of_week: Optional[int] = None
    day_of_month: Optional[int] = None

    if freq in ("weekly", "biweekly"):
        dow_label = st.selectbox("Giorno della settimana", list(DOW_LABELS.keys()), key="wiz1_dow")
        day_of_week = DOW_LABELS[dow_label]
    else:
        day_of_month = st.number_input(
            "Giorno del mese (1-28)", min_value=1, max_value=28, value=1, key="wiz1_dom"
        )

    sel_llms = st.multiselect(
        "LLM da includere nel run", LLM_OPTIONS, default=LLM_OPTIONS, key="wiz1_llms"
    )

    next_run = _calc_next_run(freq, day_of_week, int(day_of_month or 1))
    st.caption(f"Prossimo run pianificato: **{next_run.strftime('%d/%m/%Y')}**")

    c1, c2, c3 = st.columns([1, 1, 5])
    with c1:
        if st.button("← Indietro", key="s4_back"):
            _set("step", 3)
            st.rerun()
    with c2:
        if st.button("Salta e completa", key="s4_skip"):
            _finalize()
    with c3:
        if st.button("Salva e completa →", type="primary", key="s4_save",
                     disabled=not sel_llms):
            upsert_project_schedule(project_id, {
                "frequency": freq,
                "day_of_week": day_of_week,
                "day_of_month": int(day_of_month or 1) if freq == "monthly" else None,
                "llms": sel_llms,
                "is_active": True,
                "next_run_at": next_run,
            })
            _finalize()

# ===========================================================================
# STEP 5 — Completato
# ===========================================================================
elif step == 5:
    st.success(f"✅ Progetto **{_get('name')}** creato con successo!")

    col1, col2 = st.columns(2)
    with col1:
        st.metric("Progetto ID", _get("project_id") or "—")
    with col2:
        st.metric("Lingua / Paese", f"{_get('language', '')} / {_get('country', '')}")

    st.divider()
    c1, c2 = st.columns([1, 5])
    with c1:
        if st.button("Crea un altro progetto"):
            _reset()
            st.rerun()
    with c2:
        if st.button("Vai a Domande e Keyword →", type="primary"):
            _reset()
            st.switch_page("pages/3_Domande_e_Keyword.py")
