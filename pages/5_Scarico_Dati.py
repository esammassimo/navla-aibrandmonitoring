"""pages/3_Scarico_Dati.py — Manual run, partial retry, scheduling UI."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from typing import Optional

import math

import pandas as pd
import streamlit as st
from sqlalchemy import text

import pipeline as pl
from pipeline import get_run_log_path
from brand_extraction import (
    METHOD_OPTIONS, preview_extraction, run_brand_reextraction,
)
from utils import (
    LLM_GROUP,
    fetch_ai_questions,
    fetch_project_brands,
    fetch_project_schedule,
    fetch_run_workers,
    fetch_runs,
    get_cookie_manager,
    get_engine,
    render_sidebar,
    require_login,
    run_query,
    set_schedule_active,
    upsert_project_schedule,
)

cookie_manager = get_cookie_manager()
require_login(cookie_manager)
render_sidebar(cookie_manager)

is_admin = st.session_state.get("role") == "admin"
project_id: Optional[str] = st.session_state.get("project_id")

st.title("Scarico Dati")

if not project_id:
    st.info("Seleziona un progetto dalla barra laterale per iniziare.")
    st.stop()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
# LLM groups imported from utils.LLM_GROUP:
# LLM_GROUP["LLM"]         = ["ChatGPT", "Claude", "Gemini", "Perplexity"]
# LLM_GROUP["AI Features"] = ["AI Overviews", "AI Mode"]
_ITERABLE_LLMS = LLM_GROUP["LLM"]  # Only conversational LLMs support multiple iterations


def _calc_next_run(frequency: str, day_of_week: int, day_of_month: int) -> datetime:
    today = date.today()
    if frequency == "weekly":
        days_ahead = (day_of_week - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        return datetime.combine(today + timedelta(days=days_ahead), datetime.min.time())
    else:  # biweekly or monthly
        month = today.month
        year = today.year
        if today.day >= day_of_month:
            month += 1
            if month > 12:
                month = 1
                year += 1
        return datetime(year, month, min(day_of_month, 28))



# ===========================================================================
# Export helpers
# ===========================================================================

def _build_export_xlsx(project_id: str, customer_id: str) -> bytes:
    """
    Build a Google Sheets-compatible .xlsx with all historical data for the
    given project/customer. Returns the file as bytes for st.download_button.

    Sheet structure (mirrors the Apps Script template):
      1. Readme         — static description
      2. Keyword        — keywords of the project
      3. AI Questions   — active + draft questions
      4. Brand - Apps Script   — v_brand_mentions_flat
      5. Fonti - Apps Script   — v_source_mentions_flat
      6. Risposte - Apps Script — v_ai_responses_flat
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from utils import run_query, FilterState, fetch_keywords, fetch_ai_questions

    # --- Palette (matches template) ---
    HDR_BG   = "FF29282C"   # dark charcoal
    HDR_FG   = "FFF0B910"   # gold/amber
    HDR_FONT = Font(bold=True, color=HDR_FG, name="Arial", size=10)
    HDR_FILL = PatternFill("solid", fgColor=HDR_BG)
    HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
    BODY_FONT = Font(name="Arial", size=10)
    BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)

    def _write_sheet(ws, headers: list[str], rows: list[list]):
        """Write header row + data rows with consistent styling."""
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = HDR_ALIGN

        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = BODY_FONT
                cell.alignment = BODY_ALIGN

        # Auto-width (capped at 80 chars)
        for c_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(c_idx)
            cell_lens = [len(str(h))] + [
                len(str(row[c_idx - 1])) if row[c_idx - 1] is not None else 0
                for row in rows[:200]
            ]
            ws.column_dimensions[col_letter].width = min(max(cell_lens) + 2, 80)

        # Freeze header row
        ws.freeze_panes = "A2"

    # --- Fetch data ---
    filters = FilterState(
        project_id=project_id,
        customer_id=customer_id,
        date_range=None,
        llms=(),
        clusters=(),
    )

    kw_df = run_query(
        "SELECT keyword, cluster, subcluster, search_volume "
        "FROM keywords WHERE project_id = %(pid)s ORDER BY cluster, keyword",
        {"pid": project_id},
    )
    q_df = run_query(
        "SELECT aq.question, k.keyword, k.cluster, k.subcluster, k.search_volume, "
        "aq.intent, aq.tone "
        "FROM ai_questions aq "
        "LEFT JOIN keywords k ON k.id = aq.keyword_id "
        "WHERE aq.project_id = %(pid)s ORDER BY k.cluster, aq.question",
        {"pid": project_id},
    )
    brand_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, brand, position, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_brand_mentions_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm, position",
        {"pid": project_id},
    )
    source_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, url, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_source_mentions_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm",
        {"pid": project_id},
    )
    response_df = run_query(
        "SELECT date, ai_question, keyword, cluster, subcluster, volume, "
        "llm, model, response_text, "
        "(SELECT intent FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS intent, "
        "(SELECT tone   FROM ai_questions WHERE id = v.ai_question_id LIMIT 1) AS tone "
        "FROM v_ai_responses_flat v "
        "WHERE project_id = %(pid)s ORDER BY date DESC, ai_question, llm",
        {"pid": project_id},
    )

    # --- Build workbook ---
    wb = Workbook()

    # 1. Keyword
    ws_kw = wb.active
    ws_kw.title = "Keyword"
    _write_sheet(
        ws_kw,
        headers=["Keyword", "CLUSTER", "SUBCLUSTER", "Volume"],
        rows=[
            [r.keyword, r.cluster, r.subcluster, r.search_volume]
            for r in kw_df.itertuples(index=False)
        ] if not kw_df.empty else [],
    )

    # 2. AI Questions
    ws_q = wb.create_sheet("AI Questions")
    _write_sheet(
        ws_q,
        headers=["AI Questions", "Keyword", "Cluster", "Subcluster", "Volume", "Intent", "Tone"],
        rows=[
            [r.question, r.keyword, r.cluster, r.subcluster, r.search_volume, r.intent, r.tone]
            for r in q_df.itertuples(index=False)
        ] if not q_df.empty else [],
    )

    # 3. Brand - Apps Script
    ws_brand = wb.create_sheet("Brand - Apps Script")
    _write_sheet(
        ws_brand,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "Brand", "Position", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.brand, r.position, r.intent, r.tone,
            ]
            for r in brand_df.itertuples(index=False)
        ] if not brand_df.empty else [],
    )

    # 4. Fonti - Apps Script
    ws_source = wb.create_sheet("Fonti - Apps Script")
    _write_sheet(
        ws_source,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "URL", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.url, r.intent, r.tone,
            ]
            for r in source_df.itertuples(index=False)
        ] if not source_df.empty else [],
    )

    # 5. Risposte - Apps Script
    ws_resp = wb.create_sheet("Risposte - Apps Script")
    # Response text can be long — wrap text and cap row height
    ws_resp.row_dimensions[1].height = 20
    _write_sheet(
        ws_resp,
        headers=["Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "LLM", "Model", "Risposta", "Intent", "Tone"],
        rows=[
            [
                r.date.strftime("%Y-%m-%d") if hasattr(r.date, "strftime") else str(r.date),
                r.ai_question, r.keyword, r.cluster, r.subcluster, r.volume,
                r.llm, r.model, r.response_text, r.intent, r.tone,
            ]
            for r in response_df.itertuples(index=False)
        ] if not response_df.empty else [],
    )
    # Risposta column (I) — wider and wrap text
    ws_resp.column_dimensions["I"].width = 80
    for row in ws_resp.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ===========================================================================
# SECTION 1 — Avvio manuale
# ===========================================================================
st.subheader("Avvio manuale")

# Check active questions
q_df = fetch_ai_questions(project_id, status="active")
n_active = len(q_df)

if n_active == 0:
    st.warning(
        "Nessuna domanda attiva per questo progetto. "
        "Attiva almeno una domanda nella pagina **Domande e Keyword** prima di avviare un run."
    )
else:
    st.caption(f"Domande attive: **{n_active}**")

    with st.form("form_manual_run"):
        st.markdown("**LLMs to query**")
        col_llm_f, col_ai_f = st.columns(2)
        with col_llm_f:
            sel_llms_llm = st.multiselect(
                "LLM",
                options=LLM_GROUP["LLM"],
                default=LLM_GROUP["LLM"],
            )
        with col_ai_f:
            sel_llms_ai = st.multiselect(
                "AI Features",
                options=LLM_GROUP["AI Features"],
                default=LLM_GROUP["AI Features"],
            )
        selected_llms = sel_llms_llm + sel_llms_ai

        # Iterations slider — only meaningful for iterable LLMs
        iterable_selected = [l for l in selected_llms if l in _ITERABLE_LLMS] if selected_llms else []
        iterations = st.number_input(
            "Iterazioni per prompt",
            min_value=1,
            max_value=50,
            value=1,
            step=1,
            help=(
                "Numero di volte che ogni domanda viene inviata a ciascun LLM supportato "
                f"({', '.join(_ITERABLE_LLMS)}). "
                "Le iterazioni sono sequenziali. Google AIO e AI Mode vengono sempre interrogati una sola volta."
            ),
            disabled=not iterable_selected,
        )
        if iterable_selected and iterations > 1:
            st.caption(
                f"⚠️ Con **{iterations} iterazioni**, ogni domanda verrà inviata "
                f"**{iterations}×** a: {', '.join(iterable_selected)}. "
                f"Totale worker stimati: **{n_active * (len(iterable_selected) * iterations + len([l for l in selected_llms if l not in _ITERABLE_LLMS]))}**"
            )

        run_btn = st.form_submit_button("▶ Avvia run", type="primary", disabled=(n_active == 0))

    if run_btn:
        if not selected_llms:
            st.error("Please select at least one LLM.")
        else:
            total_workers = n_active * len(selected_llms)
            progress_bar = st.progress(0, text="Avvio in corso…")
            status_text = st.empty()

            def _progress_cb(done: int, total: int) -> None:
                pct = done / total if total else 0
                progress_bar.progress(pct, text=f"Worker completati: {done}/{total}")
                status_text.caption(f"In esecuzione… {done}/{total}")

            try:
                run_id = pl.start_run(
                    project_id=project_id,
                    llms=selected_llms,
                    triggered_by="manual",
                    progress_callback=_progress_cb,
                    iterations=int(iterations),
                    collect="both",
                )
                progress_bar.progress(1.0, text="Run completato.")
                status_text.empty()
                st.success(f"Run completato con successo. ID: `{run_id}`")
                fetch_runs.clear()
            except ValueError as exc:
                st.error(str(exc))
            except Exception as exc:
                st.error(f"Errore durante il run: {exc}")

# ===========================================================================
# SECTION 2 — Storico run
# ===========================================================================
st.divider()
st.subheader("Storico run")

runs_df = fetch_runs(project_id)

if runs_df.empty:
    st.info("Nessun run eseguito per questo progetto.")
else:
    display_cols = [
        "id", "started_at", "finished_at", "status",
        "triggered_by", "completed_questions", "total_questions",
    ]
    _col_config = {
        "id": st.column_config.TextColumn("Run ID"),
        "started_at": st.column_config.DatetimeColumn("Avviato", format="DD/MM/YY HH:mm"),
        "finished_at": st.column_config.DatetimeColumn("Terminato", format="DD/MM/YY HH:mm"),
        "status": st.column_config.TextColumn("Stato"),
        "triggered_by": st.column_config.TextColumn("Origine"),
        "completed_questions": st.column_config.NumberColumn("Completate"),
        "total_questions": st.column_config.NumberColumn("Totali"),
    }

    if is_admin:
        import pandas as pd

        sel_df = runs_df[display_cols].copy()
        sel_df.insert(0, "_sel", False)
        edited_runs = st.data_editor(
            sel_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "_sel": st.column_config.CheckboxColumn("Seleziona", width="small"),
                **_col_config,
            },
            disabled=display_cols,
            key="runs_selector",
        )

        selected_ids = (
            runs_df.loc[edited_runs["_sel"].values, "id"].astype(str).tolist()
        )

        confirm_key = "confirm_del_runs"
        if selected_ids:
            if not st.session_state.get(confirm_key):
                if st.button(
                    f"🗑 Elimina {len(selected_ids)} run selezionati",
                    key="del_runs_btn",
                    type="primary",
                ):
                    st.session_state[confirm_key] = selected_ids
                    st.rerun()
            else:
                ids_to_del = st.session_state[confirm_key]
                st.error(
                    f"Eliminare **{len(ids_to_del)} run** rimuoverà anche tutti i worker, "
                    "le risposte AI, le brand mentions e le source mentions associati. "
                    "Operazione non reversibile."
                )
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Sì, elimina", key="del_runs_confirm", type="primary"):
                        with get_engine().begin() as conn:
                            conn.execute(
                                text("DELETE FROM runs WHERE id::text = ANY(:ids)"),
                                {"ids": ids_to_del},
                            )
                        st.session_state.pop(confirm_key, None)
                        fetch_runs.clear()
                        st.success(f"Eliminati {len(ids_to_del)} run.")
                        st.rerun()
                with c2:
                    if st.button("Annulla", key="del_runs_cancel"):
                        st.session_state.pop(confirm_key, None)
                        st.rerun()
    else:
        st.dataframe(
            runs_df[display_cols],
            use_container_width=True,
            hide_index=True,
            column_config=_col_config,
        )

# ===========================================================================
# SECTION 2b — Run Log
# ===========================================================================
st.divider()
st.subheader("Run Log")

if runs_df.empty:
    st.info("No runs available.")
else:
    log_run_opts = {
        f"{str(row['id'])[:8]}… — {row['status']} — {str(row['started_at'])[:16]}": str(row["id"])
        for _, row in runs_df.iterrows()
    }
    selected_log_run = st.selectbox(
        "Select run to view log",
        options=list(log_run_opts.keys()),
        key="log_run_select",
    )
    selected_log_run_id = log_run_opts[selected_log_run]
    log_path = get_run_log_path(selected_log_run_id)

    if not __import__("os").path.exists(log_path):
        st.info(
            "No log file found for this run. "
            "Log files are generated starting from runs executed after this feature was deployed."
        )
    else:
        log_content = open(log_path, encoding="utf-8").read()
        n_lines = log_content.count("\n")

        # Summary badges
        n_errors   = log_content.count("[ERROR]")
        n_warnings = log_content.count("[WARNING]")
        n_info     = log_content.count("[INFO]")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Log lines",  n_lines)
        c2.metric("INFO",       n_info)
        c3.metric("WARNING",    n_warnings,  delta=None if n_warnings == 0 else f"{n_warnings}")
        c4.metric("ERROR",      n_errors,    delta=None if n_errors   == 0 else f"{n_errors}")

        # Filter options
        log_filter = st.radio(
            "Filter",
            options=["All", "INFO", "WARNING", "ERROR"],
            horizontal=True,
            key="log_filter",
            label_visibility="collapsed",
        )

        filtered_lines = [
            line for line in log_content.splitlines()
            if log_filter == "All" or f"[{log_filter}]" in line
        ]
        filtered_text = "\n".join(filtered_lines)

        # Viewer
        st.code(filtered_text or "No entries match the selected filter.", language=None)

        # Download button
        st.download_button(
            label="⬇ Download log (.txt)",
            data=log_content,
            file_name=f"run_{selected_log_run_id[:8]}.txt",
            mime="text/plain",
            key="download_log_btn",
        )

# ===========================================================================
# SECTION 3 — Retry worker falliti
# ===========================================================================
st.divider()
st.subheader("Retry worker falliti")

if runs_df.empty:
    st.info("Nessun run disponibile.")
else:
    # Only runs that have partial or failed status
    retryable = runs_df[runs_df["status"].isin(["partial", "failed"])]

    if retryable.empty:
        st.success("Nessun run con worker falliti.")
    else:
        run_opts = {
            f"{str(row['id'])[:8]}… — {row['status']} — {row['started_at']}": str(row["id"])
            for _, row in retryable.iterrows()
        }
        selected_run_label = st.selectbox(
            "Seleziona run da riprovare", list(run_opts.keys()), key="retry_run_select"
        )
        selected_run_id = run_opts[selected_run_label]

        workers_df = fetch_run_workers(selected_run_id)
        failed_workers = workers_df[workers_df["status"] == "failed"] if not workers_df.empty else workers_df

        if failed_workers.empty:
            st.info("Nessun worker fallito in questo run.")
        else:
            st.caption(f"Worker falliti: **{len(failed_workers)}**")
            st.dataframe(
                failed_workers[["question", "llm", "attempt", "error", "started_at", "finished_at"]],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "question": st.column_config.TextColumn("Domanda"),
                    "llm": st.column_config.TextColumn("LLM"),
                    "attempt": st.column_config.NumberColumn("Tentativo"),
                    "error": st.column_config.TextColumn("Errore"),
                    "started_at": st.column_config.DatetimeColumn("Avviato", format="DD/MM/YY HH:mm"),
                    "finished_at": st.column_config.DatetimeColumn("Terminato", format="DD/MM/YY HH:mm"),
                },
            )

            if st.button("🔄 Retry worker falliti", type="primary", key="retry_btn"):
                total_retry = len(failed_workers)
                progress_bar_r = st.progress(0, text="Retry in corso…")
                status_text_r = st.empty()

                def _retry_cb(done: int, total: int) -> None:
                    pct = done / total if total else 0
                    progress_bar_r.progress(pct, text=f"Retry: {done}/{total}")
                    status_text_r.caption(f"In esecuzione… {done}/{total}")

                try:
                    pl.retry_failed_workers(
                        run_id=selected_run_id,
                        progress_callback=_retry_cb,
                    )
                    progress_bar_r.progress(1.0, text="Retry completato.")
                    status_text_r.empty()
                    st.success("Retry completato.")
                    fetch_runs.clear()
                    fetch_run_workers.clear()
                    st.rerun()
                except Exception as exc:
                    st.error(f"Errore durante il retry: {exc}")

# ===========================================================================
# SECTION 3b — Brand Re-extraction
# ===========================================================================
st.divider()
st.subheader("🔄 Brand Re-extraction")
st.caption(
    "Riesegui l'estrazione brand su un run completato con un metodo diverso. "
    "Puoi testare su un campione (Preview) prima di lanciare l'estrazione completa."
)

_completed_runs = runs_df[runs_df["status"].isin(["completed", "partial"])] if not runs_df.empty else pd.DataFrame()

if _completed_runs.empty:
    st.info("Nessun run completato disponibile per la re-estrazione.")
else:
    # Selettore run
    _run_options = {
        f"{str(r.get('started_at', '?'))[:16]} — {r['status']} ({r.get('completed_questions', 0)}/{r.get('total_questions', 0)})": r["id"]
        for _, r in _completed_runs.iterrows()
    }
    _sel_run_label = st.selectbox("Run da processare", list(_run_options.keys()), key="reextract_run")
    _sel_run_id = _run_options[_sel_run_label]

    # Scelta metodo
    _sel_method = st.selectbox(
        "Metodo di estrazione",
        list(METHOD_OPTIONS.keys()),
        format_func=lambda x: METHOD_OPTIONS[x],
        key="reextract_method",
    )

    # Verifica API key
    _needs_key = None
    if "gpt-4o-mini" in _sel_method:
        _needs_key = "openai"
    elif "claude-haiku" in _sel_method:
        _needs_key = "anthropic"

    _api_ok = True
    if _needs_key:
        _has_key = bool(st.secrets.get("api_keys", {}).get(_needs_key))
        if _has_key:
            st.success(f"✅ API key `{_needs_key}` disponibile")
        else:
            st.error(f"❌ API key `{_needs_key}` mancante nei Secrets")
            _api_ok = False

        # Stima costo
        _resp_count = run_query(
            "SELECT COUNT(*) AS n FROM ai_responses "
            "WHERE run_id = %(rid)s AND response_text IS NOT NULL "
            "AND response_text != '' AND response_text NOT LIKE 'ERROR:%%' "
            "AND response_text != 'DISABLED'",
            {"rid": _sel_run_id},
        )
        _n_resp = int(_resp_count.iloc[0]["n"]) if not _resp_count.empty else 0
        st.caption(f"~{_n_resp} risposte × ~$0.001 = **~${_n_resp * 0.001:.2f}**")

    # Brand list del progetto
    _pb_df = fetch_project_brands(project_id)
    _project_brands = _pb_df.to_dict("records") if not _pb_df.empty else None
    if _project_brands:
        st.caption(f"Brand list progetto: **{len(_project_brands)}** brand configurati.")
    else:
        st.caption("⚠ Nessun brand mappato. L'estrazione userà solo pattern automatici.")

    # Check brand esistenti per resume
    _existing_df = run_query(
        "SELECT COUNT(DISTINCT ai_response_id) AS n FROM brand_mentions "
        "WHERE ai_response_id IN (SELECT id FROM ai_responses WHERE run_id = %(rid)s)",
        {"rid": _sel_run_id},
    )
    _n_existing = int(_existing_df.iloc[0]["n"]) if not _existing_df.empty else 0
    if _n_existing > 0:
        st.info(f"Questo run ha brand estratti per **{_n_existing}** risposte.")

    # ─── PREVIEW ─────────────────────────────────────────────────────────
    st.markdown("**1. Preview** — testa su un campione")

    if st.button("🔍 Preview (5 risposte)", disabled=not _api_ok, key="btn_brand_preview"):
        with st.spinner("Estrazione campione…"):
            _preview = preview_extraction(
                run_id=_sel_run_id,
                method=_sel_method,
                project_brands=_project_brands,
                sample_size=5,
            )
        if _preview:
            for pr in _preview:
                st.markdown(
                    f"**{pr['llm']}** — _{pr['question']}_\n\n"
                    f"> {pr['snippet']}…\n\n"
                    f"🏷️ **{pr['n_brands']} brand**: {', '.join(pr['brands']) if pr['brands'] else '(nessuno)'}"
                )
                st.divider()
        else:
            st.warning("Nessuna risposta valida trovata.")

    # ─── ESECUZIONE COMPLETA ─────────────────────────────────────────────
    st.markdown("**2. Esecuzione**")

    # Stop flag in session state
    if "brand_extract_stop" not in st.session_state:
        st.session_state.brand_extract_stop = False

    col_start, col_resume, col_spacer = st.columns([1, 1, 2])

    def _run_extraction(resume_mode: bool):
        st.session_state.brand_extract_stop = False
        log_lines: list[str] = []

        progress = st.progress(0, text="Avvio estrazione…")
        log_container = st.empty()
        stop_col = st.empty()

        if stop_col.button("⏹ Ferma estrazione", key=f"btn_stop_brand_{resume_mode}"):
            st.session_state.brand_extract_stop = True

        def _progress(done, total):
            progress.progress(done / max(total, 1), text=f"Estrazione: {done}/{total}")

        def _log(msg):
            log_lines.append(msg)
            log_container.code("\n".join(log_lines[-20:]), language="text")

        result = run_brand_reextraction(
            run_id=_sel_run_id,
            method=_sel_method,
            project_brands=_project_brands,
            resume=resume_mode,
            stop_flag=lambda: st.session_state.brand_extract_stop,
            progress_callback=_progress,
            log_callback=_log,
        )

        stop_col.empty()

        if result["stopped"]:
            progress.progress(
                (result["processed"] + result["skipped"]) / max(result["processed"] + result["skipped"] + 1, 1),
                text="🟡 Fermato",
            )
            st.warning(
                f"Fermato: **{result['processed']}** processate, "
                f"**{result['skipped']}** saltate, "
                f"**{result['brands_found']}** brand. "
                f"Usa **Riprendi** per continuare."
            )
        else:
            progress.progress(1.0, text="✅ Completato!")
            st.success(
                f"Completato: **{result['processed']}** processate, "
                f"**{result['skipped']}** saltate, "
                f"**{result['brands_found']}** brand, "
                f"{result['errors']} errori."
            )
            st.cache_data.clear()

    with col_start:
        _start_help = "Cancella brand esistenti e riesegue da zero" if _n_existing > 0 else "Avvia estrazione"
        if st.button("🚀 Avvia", disabled=not _api_ok, key="btn_brand_start", help=_start_help):
            _run_extraction(resume_mode=False)

    with col_resume:
        if _n_existing > 0:
            if st.button("▶ Riprendi", disabled=not _api_ok, key="btn_brand_resume",
                         help="Riparte dalle risposte non ancora processate"):
                _run_extraction(resume_mode=True)

# ===========================================================================
# SECTION 4 — Scheduling  (admin only)
# ===========================================================================
if not is_admin:
    st.stop()

st.divider()
st.subheader("Pianificazione automatica")

sched_df = fetch_project_schedule(project_id)

_FREQ_LABELS = {
    "weekly": "Settimanale",
    "biweekly": "Bimensile",
    "monthly": "Mensile",
}
_DAYS_OF_WEEK = [
    "Lunedì", "Martedì", "Mercoledì", "Giovedì",
    "Venerdì", "Sabato", "Domenica",
]

if not sched_df.empty:
    sched = sched_df.iloc[0]
    is_active: bool = bool(sched.get("is_active", False))

    col_status, col_toggle = st.columns([3, 1])
    with col_status:
        freq_label = _FREQ_LABELS.get(str(sched.get("frequency", "")), str(sched.get("frequency", "")))
        next_run = sched.get("next_run_at")
        next_str = (
            next_run.strftime("%d/%m/%Y") if hasattr(next_run, "strftime")
            else str(next_run) if next_run else "—"
        )
        status_label = "🟢 Attiva" if is_active else "🔴 Disattiva"
        st.info(
            f"**Frequenza:** {freq_label} &nbsp;|&nbsp; "
            f"**Prossimo run:** {next_str} &nbsp;|&nbsp; "
            f"**Stato:** {status_label}"
        )
    with col_toggle:
        toggle_label = "Disattiva" if is_active else "Attiva"
        if st.button(toggle_label, key="toggle_sched"):
            set_schedule_active(project_id, not is_active)
            fetch_project_schedule.clear()
            st.rerun()

st.write("**Configura pianificazione**")

with st.form("form_schedule"):
    freq = st.selectbox(
        "Frequenza",
        options=list(_FREQ_LABELS.keys()),
        format_func=lambda k: _FREQ_LABELS[k],
        index=0,
    )

    col_dow, col_dom = st.columns(2)
    with col_dow:
        day_of_week = st.selectbox(
            "Giorno della settimana (settimanale)",
            options=list(range(7)),
            format_func=lambda i: _DAYS_OF_WEEK[i],
            index=0,
            help="Usato solo per frequenza settimanale.",
        )
    with col_dom:
        day_of_month = st.number_input(
            "Giorno del mese (bimensile/mensile)",
            min_value=1, max_value=28,
            value=1,
            help="Usato per frequenza bimensile e mensile.",
        )

    st.markdown("**LLMs to query in automatic runs**")
    col_sched_a, col_sched_b = st.columns(2)
    with col_sched_a:
        sched_llms_llm = st.multiselect(
            "LLM",
            options=LLM_GROUP["LLM"],
            default=LLM_GROUP["LLM"],
        )
    with col_sched_b:
        sched_llms_ai = st.multiselect(
            "AI Features",
            options=LLM_GROUP["AI Features"],
            default=LLM_GROUP["AI Features"],
        )
    sched_llms = sched_llms_llm + sched_llms_ai

    save_sched = st.form_submit_button("Salva pianificazione", type="primary")

if save_sched:
    if not sched_llms:
        st.error("Please select at least one LLM.")
    else:
        next_run_at = _calc_next_run(freq, day_of_week, int(day_of_month))
        upsert_project_schedule(
            project_id,
            {
                "frequency": freq,
                "day_of_week": day_of_week,
                "day_of_month": int(day_of_month),
                "llms": sched_llms,
                "is_active": True,
                "next_run_at": next_run_at,
            },
        )
        fetch_project_schedule.clear()
        st.success(f"Pianificazione salvata. Prossimo run: **{next_run_at.strftime('%d/%m/%Y')}**")
        st.rerun()

# ===========================================================================
# SECTION 5 — Export Google Sheets
# ===========================================================================
st.divider()
st.subheader("Esporta dati per Google Sheets")

customer_id: Optional[str] = st.session_state.get("customer_id")

st.caption(
    "Genera un file .xlsx compatibile con il template Google Sheets contenente "
    "tutti i dati storici del progetto (keyword, domande, brand, fonti, risposte)."
)

col_exp1, col_exp2 = st.columns([3, 1])
with col_exp1:
    if runs_df.empty:
        st.info("Nessun run disponibile per questo progetto. Avvia almeno un run prima di esportare.")
    else:
        n_runs = len(runs_df)
        completed_runs = int((runs_df["status"].isin(["completed", "partial"])).sum())
        st.caption(f"Run totali: **{n_runs}** &nbsp;|&nbsp; Completati/parziali: **{completed_runs}**")

with col_exp2:
    export_disabled = runs_df.empty
    if st.button(
        "📥 Genera export",
        type="primary",
        disabled=export_disabled,
        use_container_width=True,
        key="btn_export_xlsx",
    ):
        with st.spinner("Generazione file in corso…"):
            try:
                xlsx_bytes = _build_export_xlsx(project_id, customer_id or "")
                st.session_state["export_xlsx_bytes"] = xlsx_bytes
                st.success("File generato. Clicca Download per scaricarlo.")
            except Exception as exc:
                st.error(f"Errore durante la generazione: {exc}")
                st.session_state.pop("export_xlsx_bytes", None)

if "export_xlsx_bytes" in st.session_state:
    from datetime import date as _date
    fname = f"AI_Brand_Monitor_{project_id[:8]}_{_date.today().strftime('%Y%m%d')}.xlsx"
    st.download_button(
        label="⬇ Download .xlsx",
        data=st.session_state["export_xlsx_bytes"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        key="dl_export_xlsx",
    )


# ===========================================================================
# SECTION — Import historical data
# ===========================================================================
st.divider()
st.subheader("Import historical data")
st.caption(
    "Import brand mentions, source mentions and responses from a Compass-format Excel file. "
    "The project must already have the AI Questions configured. "
    "One run is created per unique date in the file."
)

if not is_admin:
    st.info("This section is only available to administrators.")
else:
    uploaded_hist = st.file_uploader(
        "Upload Excel file (.xlsx)",
        type=["xlsx"],
        key="hist_import_upload",
    )

    if uploaded_hist:
        # ---------------------------------------------------------------
        # Preview
        # ---------------------------------------------------------------
        from openpyxl import load_workbook as _load_wb

        uploaded_hist.seek(0)
        wb_prev = _load_wb(uploaded_hist, read_only=True, data_only=True)
        uploaded_hist.seek(0)

        required_sheets = [
            "Risposte - Apps Script",
            "Brand - Apps Script",
            "Fonti - Apps Script",
        ]
        missing = [s for s in required_sheets if s not in wb_prev.sheetnames]
        if missing:
            st.error(f"Missing sheets in file: {', '.join(missing)}")
        else:
            def _load_sheet(wb, name):
                ws = wb[name]
                rows = [r for r in ws.iter_rows(values_only=True)
                        if any(c is not None for c in r)]
                if len(rows) < 2:
                    return pd.DataFrame()
                return pd.DataFrame(rows[1:], columns=rows[0])

            df_resp   = _load_sheet(wb_prev, "Risposte - Apps Script")
            df_brand  = _load_sheet(wb_prev, "Brand - Apps Script")
            df_source = _load_sheet(wb_prev, "Fonti - Apps Script")

            # Parse dates
            for df in [df_resp, df_brand, df_source]:
                if not df.empty and "Data" in df.columns:
                    df["_date"] = pd.to_datetime(df["Data"], errors="coerce").dt.date

            unique_dates = sorted(df_resp["_date"].dropna().unique()) if not df_resp.empty else []
            unique_llms  = df_resp["LLM"].dropna().unique().tolist() if not df_resp.empty else []

            col_a, col_b, col_c = st.columns(3)
            col_a.metric("Dates (runs)", len(unique_dates))
            col_b.metric("LLMs", len(unique_llms))
            col_c.metric("Responses", len(df_resp))

            c1, c2 = st.columns(2)
            with c1:
                st.caption(f"**Dates:** {', '.join(str(d) for d in unique_dates)}")
            with c2:
                st.caption(f"**LLMs:** {', '.join(unique_llms)}")

            col_d, col_e = st.columns(2)
            col_d.metric("Brand mentions", len(df_brand))
            col_e.metric("Source mentions", len(df_source))

            # Check question matching
            q_df = run_query(
                "SELECT id, question FROM ai_questions WHERE project_id = %(pid)s",
                {"pid": project_id},
            )
            q_map = {
                str(r["question"]).strip().lower(): str(r["id"])
                for _, r in q_df.iterrows()
            } if not q_df.empty else {}

            resp_questions = set(
                str(r).strip().lower()
                for r in df_resp["AI Questions"].dropna()
            ) if not df_resp.empty else set()
            matched   = resp_questions & set(q_map.keys())
            unmatched = resp_questions - set(q_map.keys())

            if unmatched:
                st.warning(
                    f"**{len(matched)}** of {len(resp_questions)} questions matched in DB. "
                    f"**{len(unmatched)}** not found — responses will be saved with `ai_question_id = NULL`."
                )
            else:
                st.success(f"✅ All {len(matched)} questions matched in DB.")

            # ---------------------------------------------------------------
            # Import button
            # ---------------------------------------------------------------
            already_imported = run_query(
                "SELECT run_date FROM runs WHERE project_id = %(pid)s "
                "AND triggered_by = 'manual' AND status = 'completed'",
                {"pid": project_id},
            )
            already_dates = set()
            if not already_imported.empty and "run_date" in already_imported.columns:
                already_dates = set(
                    pd.to_datetime(already_imported["run_date"]).dt.date.tolist()
                )

            overlap = [d for d in unique_dates if d in already_dates]
            if overlap:
                st.warning(
                    f"⚠ Dates already imported: {', '.join(str(d) for d in overlap)}. "
                    "Re-importing will create duplicate runs for those dates."
                )

            if st.button("⬆ Import historical data", type="primary",
                         key="btn_hist_import"):
                uploaded_hist.seek(0)
                wb_imp = _load_wb(uploaded_hist, read_only=True, data_only=True)

                def _load_sheet_imp(name):
                    ws = wb_imp[name]
                    rows = [r for r in ws.iter_rows(values_only=True)
                            if any(c is not None for c in r)]
                    if len(rows) < 2:
                        return pd.DataFrame()
                    df = pd.DataFrame(rows[1:], columns=rows[0])
                    if "Data" in df.columns:
                        df["_date"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
                    return df

                dr = _load_sheet_imp("Risposte - Apps Script")
                db = _load_sheet_imp("Brand - Apps Script")
                ds = _load_sheet_imp("Fonti - Apps Script")

                total_runs = 0
                total_resp = 0
                total_brand = 0
                total_source = 0
                errors = []

                progress = st.progress(0, text="Starting import…")
                run_dates = sorted(dr["_date"].dropna().unique())

                for i, run_date in enumerate(run_dates):
                    progress.progress(
                        (i) / len(run_dates),
                        text=f"Importing {run_date} ({i+1}/{len(run_dates)})…"
                    )
                    resp_group  = dr[dr["_date"] == run_date]
                    brand_group = db[db["_date"] == run_date] if not db.empty else pd.DataFrame()
                    source_group = ds[ds["_date"] == run_date] if not ds.empty else pd.DataFrame()

                    llms_in_run = resp_group["LLM"].dropna().unique().tolist()
                    n_q = len(resp_group)

                    # Create run
                    try:
                        with get_engine().begin() as conn:
                            row = conn.execute(
                                text(
                                    "INSERT INTO runs "
                                    "(project_id, started_at, finished_at, status, "
                                    " triggered_by, llms, total_questions, completed_questions) "
                                    "VALUES (:pid, :started, :finished, 'completed', "
                                    "        'manual', :llms, :total, :total) "
                                    "RETURNING id"
                                ),
                                {
                                    "pid":     project_id,
                                    "started": datetime.combine(run_date, datetime.min.time()),
                                    "finished": datetime.combine(run_date, datetime.min.time()),
                                    "llms":    llms_in_run,
                                    "total":   n_q,
                                },
                            ).fetchone()
                        run_id = str(row[0])
                        total_runs += 1
                    except Exception as e:
                        errors.append(f"{run_date}: run creation failed — {e}")
                        continue

                    # Insert responses + brand + source mentions
                    for _, rrow in resp_group.iterrows():
                        q_text  = str(rrow.get("AI Questions", "")).strip()
                        llm     = str(rrow.get("LLM", "")).strip()
                        model   = str(rrow.get("Model", "")).strip()
                        resp_text = str(rrow.get("Risposta", "")).strip() or None
                        q_id    = q_map.get(q_text.lower())

                        try:
                            with get_engine().begin() as conn:
                                ar_row = conn.execute(
                                    text(
                                        "INSERT INTO ai_responses "
                                        "(run_id, ai_question_id, llm, model, "
                                        " response_text, run_date) "
                                        "VALUES (:rid, :qid, :llm, :model, :text, :rdate) "
                                        "RETURNING id"
                                    ),
                                    {
                                        "rid":   run_id,
                                        "qid":   q_id,
                                        "llm":   llm,
                                        "model": model,
                                        "text":  resp_text,
                                        "rdate": run_date,
                                    },
                                ).fetchone()
                            response_id = str(ar_row[0])
                            total_resp += 1
                        except Exception as e:
                            errors.append(f"{run_date} / {q_text[:40]}: response failed — {e}")
                            continue

                        # Brand mentions
                        b_match = brand_group[
                            (brand_group["AI Questions"].astype(str).str.strip() == q_text) &
                            (brand_group["LLM"].astype(str).str.strip() == llm)
                        ] if not brand_group.empty else pd.DataFrame()

                        if not b_match.empty:
                            try:
                                with get_engine().begin() as conn:
                                    for _, br in b_match.iterrows():
                                        brand_name = str(br.get("Brand", "")).strip()
                                        if not brand_name:
                                            continue
                                        pos = br.get("Position")
                                        pos_int = (
                                            int(pos) if pos is not None and
                                            not (isinstance(pos, float) and math.isnan(pos))
                                            else None
                                        )
                                        conn.execute(
                                            text("INSERT INTO brand_mentions "
                                                 "(ai_response_id, brand_name, position) "
                                                 "VALUES (:rid, :brand, :pos)"),
                                            {"rid": response_id, "brand": brand_name, "pos": pos_int},
                                        )
                                        total_brand += 1
                            except Exception as e:
                                errors.append(f"brand mentions {run_date}/{q_text[:30]}: {e}")

                        # Source mentions
                        s_match = source_group[
                            (source_group["AI Questions"].astype(str).str.strip() == q_text) &
                            (source_group["LLM"].astype(str).str.strip() == llm)
                        ] if not source_group.empty else pd.DataFrame()

                        if not s_match.empty:
                            try:
                                with get_engine().begin() as conn:
                                    for _, sr in s_match.iterrows():
                                        url = str(sr.get("URL", "")).strip()
                                        if not url:
                                            continue
                                        conn.execute(
                                            text("INSERT INTO source_mentions "
                                                 "(ai_response_id, url) "
                                                 "VALUES (:rid, :url)"),
                                            {"rid": response_id, "url": url},
                                        )
                                        total_source += 1
                            except Exception as e:
                                errors.append(f"source mentions {run_date}/{q_text[:30]}: {e}")

                progress.progress(1.0, text="Import complete.")
                st.cache_data.clear()
                fetch_runs.clear()

                if errors:
                    with st.expander(f"⚠ {len(errors)} error(s) during import"):
                        for err in errors:
                            st.text(err)

                st.success(
                    f"✅ Import complete: "
                    f"**{total_runs}** run(s) · "
                    f"**{total_resp}** responses · "
                    f"**{total_brand}** brand mentions · "
                    f"**{total_source}** source mentions"
                )
                st.rerun()
